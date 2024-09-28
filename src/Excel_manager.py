import asyncio
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelManager:
    """Класс-менеджер для работы с Excel."""

    def __init__(self) -> None:
        """Инициализация менеджера excel."""

        self.workbook = None
        self.sheet = None
        self.file_name = None

    async def create_excel_file(self) -> None:
        """Асинхронно создает новый файл Excel с заголовками столбцов."""

        self.file_name = self.create_filename()

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Weather Data"

        # Заголовки для столбцов
        headers = [
            "№",
            "Широта",
            "Долгота",
            "Часовой пояс",
            "Смещение часового пояса",
            "Дата и время запроса данных",
            "Дата и время измерения погоды",
            "Температура воздуха (°C)",
            "Количество осадков (мм)",
            "Тип осадков",
            "Атмосферное давление (мм рт.ст.)",
            "Скорость ветра (m/s)",
            "Направление ветра",
        ]

        # Заполняем первую строку заголовками
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            self.sheet[f"{column_letter}1"] = header

        await asyncio.sleep(0)

        self.workbook.save(self.file_name)

    async def add_weather_data_to_excel(self, weather_data: list[dict]) -> None:
        """Асинхронно добавляет данные о погоде в Excel-файл.

        Args:
            weather_data (list[dict]): список данных о погоде для записи в файл.
        """

        if not self.workbook or not self.sheet:
            raise ValueError("Файл Excel не был создан.")

        # Стартовая строка для добавления данных
        start_row = self.sheet.max_row + 1

        for row_num, data in enumerate(weather_data, start=start_row):
            self.sheet[f"A{row_num}"] = data.get("id")
            self.sheet[f"B{row_num}"] = data.get("latitude")
            self.sheet[f"C{row_num}"] = data.get("longitude")
            self.sheet[f"D{row_num}"] = data.get("timezone")
            self.sheet[f"E{row_num}"] = data.get("utc_offset_seconds")
            self.sheet[f"F{row_num}"] = (
                data.get("datetime_request").strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(data.get("datetime_request"), datetime)
                else data.get("datetime_request")
            )
            self.sheet[f"G{row_num}"] = (
                data.get("datetime_weather").strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(data.get("datetime_weather"), datetime)
                else data.get("datetime_weather")
            )
            self.sheet[f"H{row_num}"] = data.get("temperature_2m")
            self.sheet[f"I{row_num}"] = data.get("precipitation")
            self.sheet[f"J{row_num}"] = data.get("type_precipitation")
            self.sheet[f"K{row_num}"] = data.get("pressure_msl")
            self.sheet[f"L{row_num}"] = data.get("wind_speed_10m")
            self.sheet[f"M{row_num}"] = data.get("wind_direction_10m")

        # Сохраняем изменения в файле
        self.workbook.save(self.file_name)
        await asyncio.sleep(0)

    @staticmethod
    def create_filename() -> str:
        """Создаёт имя файла с текущей датой и временем.

        Returns:
            str: Имя файла с меткой времени.
        """

        return f"weather_data_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
